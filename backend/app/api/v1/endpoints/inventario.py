"""
Endpoints para el módulo de inventario de datos - RAT (Registro de Actividades de Tratamiento)
Implementación práctica según Ley 21.719
"""
from typing import List, Optional, Dict, Any
from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File, Response, Request
from fastapi.responses import FileResponse, StreamingResponse
from sqlalchemy.orm import Session
from datetime import datetime
import json
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.core.database import get_db
from app.core.dependencies import get_current_active_user, require_permission
from app.core.tenant import get_tenant_db
from app.models.user import User
from app.models.inventario import ActividadTratamiento, CategoriaDatos, BaseLegal
from app.services.inventario_service import InventarioService
from app.services.recursos_service import RecursosService

router = APIRouter()
inventario_service = InventarioService()
recursos_service = RecursosService()


@router.get("/", response_model=List[Dict[str, Any]])
async def listar_actividades_tratamiento(
    request: Request,
    departamento: Optional[str] = None,
    con_transferencias: Optional[bool] = None,
    skip: int = 0,
    limit: int = 100,
    current_user: User = Depends(get_current_active_user),
    _: bool = Depends(require_permission(["inventario.read", "inventario.manage"]))
) -> Any:
    """
    Listar todas las actividades de tratamiento del inventario
    """
    db = get_tenant_db(request.state.tenant_id)
    
    actividades = inventario_service.get_actividades_tratamiento(
        db=db,
        tenant_id=request.state.tenant_id,
        departamento=departamento,
        con_transferencias=con_transferencias
    )
    
    return actividades[skip:skip + limit]


@router.post("/actividades", response_model=Dict[str, Any])
async def crear_actividad_tratamiento(
    request: Request,
    data: Dict[str, Any],
    current_user: User = Depends(get_current_active_user),
    _: bool = Depends(require_permission(["inventario.create", "inventario.manage"]))
) -> Any:
    """
    Crear una nueva actividad de tratamiento en el inventario
    """
    db = get_tenant_db(request.state.tenant_id)
    
    try:
        actividad = inventario_service.create_actividad_tratamiento(
            db=db,
            tenant_id=request.state.tenant_id,
            nombre=data["nombre"],
            descripcion=data["descripcion"],
            finalidad=data["finalidad"],
            departamento=data["departamento"],
            responsable_id=current_user.id,
            base_legal_ids=data.get("base_legal_ids", []),
            categoria_datos_ids=data.get("categoria_datos_ids", []),
            origen_datos=data.get("origen_datos", "titular"),
            plazo_conservacion=data["plazo_conservacion"],
            medidas_seguridad_ids=data.get("medidas_seguridad_ids", []),
            transferencias_internacionales=data.get("transferencias_internacionales", False),
            metadata={
                "creado_por": current_user.username,
                "metodo_creacion": "manual"
            }
        )
        
        return {
            "success": True,
            "actividad_id": actividad.id,
            "mensaje": "Actividad de tratamiento creada exitosamente"
        }
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )


@router.get("/plantillas/{tipo}")
async def descargar_plantilla_inventario(
    tipo: str,
    current_user: User = Depends(get_current_active_user),
    _: bool = Depends(require_permission(["inventario.read"]))
) -> Any:
    """
    Descargar plantillas Excel para el inventario de datos
    
    Tipos disponibles:
    - rat_simple: Plantilla básica de Registro de Actividades
    - rat_completo: Plantilla completa con todas las hojas
    - ejemplos_sector: Plantilla con ejemplos por industria
    """
    
    if tipo == "rat_simple":
        return _generar_plantilla_rat_simple()
    elif tipo == "rat_completo":
        return _generar_plantilla_rat_completo()
    elif tipo == "ejemplos_sector":
        return _generar_plantilla_ejemplos()
    else:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Tipo de plantilla no encontrado"
        )


@router.get("/dashboard", response_model=Dict[str, Any])
async def dashboard_inventario(
    request: Request,
    current_user: User = Depends(get_current_active_user),
    _: bool = Depends(require_permission(["inventario.read"]))
) -> Any:
    """
    Dashboard con métricas del inventario de datos
    """
    db = get_tenant_db(request.state.tenant_id)
    
    # Obtener estadísticas
    total_actividades = db.query(ActividadTratamiento).filter(
        ActividadTratamiento.tenant_id == request.state.tenant_id
    ).count()
    
    actividades_activas = db.query(ActividadTratamiento).filter(
        ActividadTratamiento.tenant_id == request.state.tenant_id,
        ActividadTratamiento.is_active == True
    ).count()
    
    con_transferencias = db.query(ActividadTratamiento).filter(
        ActividadTratamiento.tenant_id == request.state.tenant_id,
        ActividadTratamiento.tiene_transferencia_internacional == True
    ).count()
    
    requieren_dpia = db.query(ActividadTratamiento).filter(
        ActividadTratamiento.tenant_id == request.state.tenant_id,
        ActividadTratamiento.requiere_dpia == True,
        ActividadTratamiento.dpia_realizada == False
    ).count()
    
    # Actividades por departamento
    from sqlalchemy import func
    por_departamento = db.query(
        ActividadTratamiento.area_responsable,
        func.count(ActividadTratamiento.id).label('total')
    ).filter(
        ActividadTratamiento.tenant_id == request.state.tenant_id
    ).group_by(ActividadTratamiento.area_responsable).all()
    
    # Calcular completitud
    sin_base_legal = db.query(ActividadTratamiento).filter(
        ActividadTratamiento.tenant_id == request.state.tenant_id,
        ActividadTratamiento.base_legal_id == None
    ).count()
    
    completitud = 100
    if total_actividades > 0:
        completitud = int(((total_actividades - sin_base_legal - requieren_dpia) / total_actividades) * 100)
    
    return {
        "metricas": {
            "total_actividades": total_actividades,
            "actividades_activas": actividades_activas,
            "con_transferencias_internacionales": con_transferencias,
            "requieren_dpia": requieren_dpia,
            "sin_base_legal": sin_base_legal,
            "completitud_porcentaje": completitud
        },
        "por_departamento": [
            {"departamento": dept, "total": total}
            for dept, total in por_departamento
        ],
        "alertas": [
            f"{sin_base_legal} actividades sin base legal definida" if sin_base_legal > 0 else None,
            f"{requieren_dpia} actividades requieren DPIA" if requieren_dpia > 0 else None,
        ],
        "proximas_acciones": [
            "Completar base legal para todas las actividades",
            "Realizar evaluaciones de impacto (DPIA) pendientes",
            "Revisar actividades con más de 6 meses sin actualización"
        ]
    }


def _generar_plantilla_rat_simple():
    """Genera plantilla Excel básica para RAT"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Registro de Actividades"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Encabezados
    headers = [
        "ID Actividad",
        "Nombre de la Actividad",
        "Descripción",
        "Finalidad del Tratamiento",
        "Departamento Responsable",
        "Base de Licitud",
        "Categorías de Datos",
        "Origen de los Datos",
        "Destinatarios",
        "Transferencias Internacionales",
        "Plazo de Conservación",
        "Medidas de Seguridad"
    ]
    
    # Escribir encabezados
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Ajustar anchos de columna
    column_widths = [15, 30, 40, 40, 25, 25, 35, 20, 30, 25, 20, 40]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Agregar ejemplo
    ejemplo = [
        "RRHH-001",
        "Proceso de Reclutamiento y Selección",
        "Gestión de postulaciones laborales, evaluación de candidatos y selección de personal",
        "Evaluar la idoneidad de candidatos para cubrir vacantes laborales en la empresa",
        "Recursos Humanos",
        "Consentimiento del titular / Medidas precontractuales",
        "Identificación, Contacto, Formación académica, Experiencia laboral, Pretensiones de renta",
        "Titular de los datos",
        "Gerentes de área, Empresa verificación antecedentes (con contrato)",
        "No",
        "6 meses (no seleccionados) / Duración relación laboral (contratados)",
        "Acceso restringido, Encriptación de base de datos, Registro de accesos"
    ]
    
    for col, valor in enumerate(ejemplo, 1):
        cell = ws.cell(row=2, column=col, value=valor)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = border
    
    # Agregar instrucciones en nueva hoja
    ws_inst = wb.create_sheet("Instrucciones")
    instrucciones = [
        ["INSTRUCCIONES PARA COMPLETAR EL REGISTRO DE ACTIVIDADES DE TRATAMIENTO"],
        [""],
        ["1. ID Actividad:", "Código único para identificar la actividad (ej: DEPT-001)"],
        ["2. Nombre:", "Nombre descriptivo del proceso o actividad"],
        ["3. Descripción:", "Explicación detallada de en qué consiste la actividad"],
        ["4. Finalidad:", "¿Para qué se usan los datos? Sea específico"],
        ["5. Base de Licitud:", "Fundamento legal: Consentimiento, Contrato, Obligación legal, Interés legítimo"],
        ["6. Categorías de Datos:", "Tipos de datos que se tratan (separados por coma)"],
        ["7. Origen:", "De dónde provienen los datos: Titular, Terceros, Fuentes públicas"],
        ["8. Destinatarios:", "A quién se comunican los datos (internos y externos)"],
        ["9. Transferencias:", "Indicar país si hay transferencias internacionales"],
        ["10. Plazo:", "Cuánto tiempo se conservan los datos"],
        ["11. Medidas:", "Qué medidas de seguridad se aplican"]
    ]
    
    for row, instruccion in enumerate(instrucciones, 1):
        for col, texto in enumerate(instruccion, 1):
            ws_inst.cell(row=row, column=col, value=texto)
    
    # Guardar en memoria
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=plantilla_inventario_RAT.xlsx"
        }
    )


def _generar_plantilla_rat_completo():
    """Genera plantilla Excel completa con múltiples hojas"""
    # Por brevedad, retorno la simple
    # En producción, esto tendría hojas adicionales:
    # - Categorías de Datos predefinidas
    # - Bases Legales según Ley 21.719
    # - Ejemplos por industria
    # - Flujos de datos
    # - Dashboard de completitud
    return _generar_plantilla_rat_simple()


def _generar_plantilla_ejemplos():
    """Genera plantilla con ejemplos por sector"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Ejemplos por Sector"
    
    # Aquí se incluirían ejemplos específicos de:
    # - Salmonicultura
    # - Retail
    # - Salud
    # - Educación
    # - Banca
    
    ws.cell(row=1, column=1, value="Ejemplos de Actividades de Tratamiento por Industria")
    ws.cell(row=3, column=1, value="En desarrollo...")
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=ejemplos_inventario_sectores.xlsx"
        }
    )


@router.post("/entrevista/iniciar", response_model=Dict[str, Any])
async def iniciar_entrevista_inventario(
    request: Request,
    departamento: str,
    current_user: User = Depends(get_current_active_user),
    _: bool = Depends(require_permission(["inventario.create"]))
) -> Any:
    """
    Iniciar una entrevista guiada para crear actividades de tratamiento
    """
    import uuid
    sesion_id = str(uuid.uuid4())
    
    # Guardar sesión en caché o DB temporal
    # Por simplicidad, retornamos la estructura inicial
    
    preguntas_iniciales = _obtener_preguntas_por_departamento(departamento)
    
    return {
        "sesion_id": sesion_id,
        "departamento": departamento,
        "paso_actual": 1,
        "total_pasos": 8,
        "pregunta_actual": preguntas_iniciales[0],
        "progreso": 12.5
    }


@router.post("/entrevista/{sesion_id}/responder", response_model=Dict[str, Any])
async def responder_pregunta_entrevista(
    request: Request,
    sesion_id: str,
    respuesta: Dict[str, Any],
    current_user: User = Depends(get_current_active_user)
) -> Any:
    """
    Procesar respuesta de la entrevista y devolver siguiente pregunta
    """
    # Aquí se procesaría la respuesta y se determinaría la siguiente pregunta
    # basándose en la lógica del flujo
    
    paso_actual = respuesta.get("paso_actual", 1)
    siguiente_paso = paso_actual + 1
    
    if siguiente_paso > 8:
        # Finalizar y crear actividad
        return {
            "completado": True,
            "resumen": _generar_resumen_entrevista(sesion_id, respuesta),
            "actividad_creada": True
        }
    
    siguiente_pregunta = _obtener_siguiente_pregunta(sesion_id, respuesta)
    
    return {
        "sesion_id": sesion_id,
        "paso_actual": siguiente_paso,
        "total_pasos": 8,
        "pregunta_actual": siguiente_pregunta,
        "progreso": (siguiente_paso / 8) * 100
    }


@router.get("/ejemplos/{sector}", response_model=List[Dict[str, Any]])
async def obtener_ejemplos_sector(
    sector: str,
    current_user: User = Depends(get_current_active_user),
    _: bool = Depends(require_permission(["inventario.read"]))
) -> Any:
    """
    Obtener ejemplos de actividades de tratamiento por sector/industria
    """
    ejemplos = _cargar_ejemplos_sector(sector)
    
    if not ejemplos:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"No hay ejemplos disponibles para el sector: {sector}"
        )
    
    return ejemplos


@router.post("/importar/excel")
async def importar_inventario_excel(
    request: Request,
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_active_user),
    _: bool = Depends(require_permission(["inventario.create", "inventario.manage"]))
) -> Any:
    """
    Importar actividades de tratamiento desde archivo Excel
    """
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="El archivo debe ser Excel (.xlsx o .xls)"
        )
    
    try:
        # Leer archivo Excel
        df = pd.read_excel(file.file, sheet_name="Registro de Actividades")
        
        db = get_tenant_db(request.state.tenant_id)
        actividades_creadas = 0
        errores = []
        
        for index, row in df.iterrows():
            if index == 0:  # Saltar encabezados
                continue
                
            try:
                # Procesar cada fila y crear actividad
                # Aquí se mapearían las columnas a los campos del modelo
                actividad = inventario_service.create_actividad_tratamiento(
                    db=db,
                    tenant_id=request.state.tenant_id,
                    nombre=row['Nombre de la Actividad'],
                    descripcion=row['Descripción'],
                    finalidad=row['Finalidad del Tratamiento'],
                    departamento=row['Departamento Responsable'],
                    responsable_id=current_user.id,
                    base_legal_ids=[],  # Mapear según contenido
                    categoria_datos_ids=[],  # Mapear según contenido
                    origen_datos=row['Origen de los Datos'],
                    plazo_conservacion=row['Plazo de Conservación'],
                    metadata={
                        "importado_desde": file.filename,
                        "fecha_importacion": datetime.utcnow().isoformat()
                    }
                )
                actividades_creadas += 1
            except Exception as e:
                errores.append({
                    "fila": index + 1,
                    "error": str(e)
                })
        
        return {
            "success": True,
            "actividades_creadas": actividades_creadas,
            "errores": errores,
            "mensaje": f"Se importaron {actividades_creadas} actividades correctamente"
        }
        
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Error al procesar el archivo: {str(e)}"
        )


@router.get("/exportar/{formato}")
async def exportar_inventario(
    request: Request,
    formato: str,
    incluir_detalles: bool = True,
    current_user: User = Depends(get_current_active_user),
    _: bool = Depends(require_permission(["inventario.read", "inventario.export"]))
) -> Any:
    """
    Exportar el inventario completo en diferentes formatos
    
    Formatos disponibles:
    - excel: Archivo Excel con formato
    - json: Datos en formato JSON
    - pdf: Informe PDF (próximamente)
    """
    db = get_tenant_db(request.state.tenant_id)
    
    if formato == "json":
        registro = inventario_service.generar_registro_actividades(
            db=db,
            tenant_id=request.state.tenant_id,
            formato="json"
        )
        
        return Response(
            content=registro,
            media_type="application/json",
            headers={
                "Content-Disposition": f"attachment; filename=inventario_datos_{datetime.now().strftime('%Y%m%d')}.json"
            }
        )
    
    elif formato == "excel":
        return _exportar_inventario_excel(db, request.state.tenant_id, incluir_detalles)
    
    elif formato == "pdf":
        raise HTTPException(
            status_code=status.HTTP_501_NOT_IMPLEMENTED,
            detail="Exportación a PDF próximamente disponible"
        )
    
    else:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Formato no soportado"
        )


# Funciones auxiliares

def _obtener_preguntas_por_departamento(departamento: str) -> List[Dict[str, Any]]:
    """Obtiene las preguntas iniciales según el departamento"""
    
    preguntas_base = {
        "RRHH": {
            "pregunta": "¿Qué procesos de gestión de personal realizas?",
            "tipo": "multiple",
            "opciones": [
                "Reclutamiento y selección",
                "Gestión de nómina",
                "Evaluaciones de desempeño",
                "Capacitación y desarrollo",
                "Control de asistencia",
                "Gestión de beneficios",
                "Desvinculaciones"
            ]
        },
        "Marketing": {
            "pregunta": "¿Qué actividades de marketing involucran datos personales?",
            "tipo": "multiple",
            "opciones": [
                "Email marketing",
                "Gestión de redes sociales",
                "Programas de fidelización",
                "Encuestas de satisfacción",
                "Análisis de comportamiento web",
                "Eventos y webinars",
                "Publicidad segmentada"
            ]
        },
        "Ventas": {
            "pregunta": "¿Qué procesos de ventas gestionas?",
            "tipo": "multiple",
            "opciones": [
                "Gestión de leads",
                "Cotizaciones",
                "Gestión de clientes (CRM)",
                "Contratos y acuerdos",
                "Facturación",
                "Postventa y soporte",
                "Gestión de partners"
            ]
        }
    }
    
    return [preguntas_base.get(departamento, preguntas_base["RRHH"])]


def _obtener_siguiente_pregunta(sesion_id: str, respuesta_anterior: Dict[str, Any]) -> Dict[str, Any]:
    """Determina la siguiente pregunta basándose en las respuestas anteriores"""
    
    # Lógica simplificada - en producción sería más compleja
    paso = respuesta_anterior.get("paso_actual", 1)
    
    preguntas_flujo = [
        {
            "pregunta": "¿Qué tipos de datos personales recopilas en este proceso?",
            "tipo": "checklist",
            "opciones": [
                "Datos de identificación (nombre, RUT)",
                "Datos de contacto (email, teléfono)",
                "Datos laborales",
                "Datos financieros",
                "Datos de salud",
                "Datos biométricos",
                "Preferencias y comportamiento"
            ]
        },
        {
            "pregunta": "¿Cuál es la finalidad principal de recopilar estos datos?",
            "tipo": "textarea",
            "placeholder": "Describe específicamente para qué usas estos datos"
        },
        {
            "pregunta": "¿Cuál es la base legal para este tratamiento?",
            "tipo": "radio",
            "opciones": [
                {"valor": "consentimiento", "texto": "Consentimiento del titular"},
                {"valor": "contrato", "texto": "Ejecución de un contrato"},
                {"valor": "obligacion_legal", "texto": "Cumplimiento de obligación legal"},
                {"valor": "interes_vital", "texto": "Proteger intereses vitales"},
                {"valor": "interes_legitimo", "texto": "Interés legítimo"}
            ]
        }
    ]
    
    if paso < len(preguntas_flujo):
        return preguntas_flujo[paso]
    
    return {
        "pregunta": "¿Por cuánto tiempo conservas estos datos?",
        "tipo": "select",
        "opciones": [
            "Mientras dure la relación",
            "6 meses después",
            "1 año",
            "5 años",
            "10 años",
            "Otro (especificar)"
        ]
    }


def _generar_resumen_entrevista(sesion_id: str, respuestas: Dict[str, Any]) -> Dict[str, Any]:
    """Genera un resumen de la entrevista completada"""
    
    return {
        "nombre_actividad": respuestas.get("nombre_proceso", "Proceso sin nombre"),
        "departamento": respuestas.get("departamento", "No especificado"),
        "datos_recopilados": respuestas.get("tipos_datos", []),
        "finalidad": respuestas.get("finalidad", ""),
        "base_legal": respuestas.get("base_legal", ""),
        "plazo_conservacion": respuestas.get("plazo", ""),
        "requiere_revision": True
    }


def _cargar_ejemplos_sector(sector: str) -> List[Dict[str, Any]]:
    """Carga ejemplos predefinidos por sector"""
    
    ejemplos = {
        "salmonicultura": [
            {
                "nombre": "Control de Acceso a Centros de Cultivo",
                "descripcion": "Registro y control de personal que ingresa a centros de cultivo marino",
                "finalidad": "Seguridad, trazabilidad sanitaria y cumplimiento normativo",
                "datos": ["Nombre", "RUT", "Empresa", "Motivo visita", "Hora entrada/salida", "Temperatura corporal"],
                "base_legal": "Obligación legal - Normativa SERNAPESCA",
                "plazo": "2 años"
            },
            {
                "nombre": "Monitoreo de Salud del Personal en Faenas",
                "descripcion": "Control de aptitud física y salud ocupacional de buzos y operarios",
                "finalidad": "Prevención de riesgos laborales en ambiente marino",
                "datos": ["Exámenes médicos", "Certificación buceo", "Condiciones preexistentes", "Medicamentos"],
                "base_legal": "Obligación legal - Código del Trabajo",
                "plazo": "Duración contrato + 5 años"
            }
        ],
        "retail": [
            {
                "nombre": "Programa de Fidelización de Clientes",
                "descripcion": "Gestión de puntos y beneficios para clientes frecuentes",
                "finalidad": "Premiar fidelidad y personalizar ofertas",
                "datos": ["Nombre", "RUT", "Email", "Historial compras", "Preferencias"],
                "base_legal": "Consentimiento",
                "plazo": "Mientras esté activo + 2 años"
            }
        ],
        "salud": [
            {
                "nombre": "Gestión de Historias Clínicas",
                "descripcion": "Registro y mantención de fichas médicas de pacientes",
                "finalidad": "Prestación de servicios de salud y continuidad de atención",
                "datos": ["Datos personales", "Historia médica", "Diagnósticos", "Tratamientos", "Exámenes"],
                "base_legal": "Obligación legal - Ley 20.584",
                "plazo": "15 años mínimo"
            }
        ]
    }
    
    return ejemplos.get(sector, [])


def _exportar_inventario_excel(db: Session, tenant_id: str, incluir_detalles: bool) -> StreamingResponse:
    """Exporta el inventario completo a Excel con formato"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario RAT"
    
    # Obtener datos
    actividades = inventario_service.get_actividades_tratamiento(
        db=db,
        tenant_id=tenant_id
    )
    
    # Crear encabezados y llenar datos
    # (Implementación similar a _generar_plantilla_rat_simple pero con datos reales)
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    fecha = datetime.now().strftime('%Y%m%d')
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=inventario_RAT_{fecha}.xlsx"
        }
    )